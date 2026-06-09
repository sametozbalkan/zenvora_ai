import fitz  # PyMuPDF
import pytesseract
from PIL import Image, ImageOps
from pathlib import Path
from docx import Document
from contextlib import closing
import io
import shutil
import os

tesseract_path = shutil.which("tesseract") or os.getenv("TESSERACT_PATH", "tesseract")
pytesseract.pytesseract.tesseract_cmd = tesseract_path


def parse_pdf(file_bytes: bytes) -> str:
    # closing ile fitz Document handle'ı garanti kapansın
    with closing(fitz.open(stream=file_bytes, filetype="pdf")) as doc:
        return "\n".join(
            page.get_text("text", flags=fitz.TEXT_PRESERVE_WHITESPACE)
            for page in doc
        )


def parse_image(file_bytes: bytes) -> str:
    image = Image.open(io.BytesIO(file_bytes))
    # Telefon kameralarıyla çekilen görseller EXIF "Orientation" etiketi ile
    # 90/180/270° döndürülmüş gelir; transpose etmezsek OCR ters/yan okur.
    image = ImageOps.exif_transpose(image)
    return pytesseract.image_to_string(image, lang="tur+eng")


def parse_docx(file_bytes: bytes) -> str:
    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


MAX_XLSX_ROWS_PER_SHEET = 200


def parse_xlsx(file_bytes: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    out: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.append(f"## Sayfa: {sheet_name}")
        # max_row None olabilir; güvenli okuma
        total = ws.max_row or 0
        if total > MAX_XLSX_ROWS_PER_SHEET:
            out.append(
                f"[Çok büyük tablo, ilk {MAX_XLSX_ROWS_PER_SHEET} satır işlendi]"
            )
        # iter_rows + max_row sınırı: bellek dostu
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx > MAX_XLSX_ROWS_PER_SHEET:
                break
            line = " | ".join(str(c) if c is not None else "" for c in row)
            if line.strip(" |"):
                out.append(line)
    return "\n".join(out)


# Magic-byte imzaları — uzantıyla içerik tutarlılığını doğrular.
# Birden fazla varyant olabilen tipler için tuple listesi.
_MAGIC_SIGNATURES: dict[str, tuple[bytes, ...]] = {
    ".pdf":  (b"%PDF-",),
    ".png":  (b"\x89PNG\r\n\x1a\n",),
    ".jpg":  (b"\xff\xd8\xff",),
    ".jpeg": (b"\xff\xd8\xff",),
    ".bmp":  (b"BM",),
    ".tiff": (b"II*\x00", b"MM\x00*"),
    # docx/xlsx ZIP container (PK\x03\x04) ya da boş arşiv
    ".docx": (b"PK\x03\x04", b"PK\x05\x06"),
    ".xlsx": (b"PK\x03\x04", b"PK\x05\x06"),
}


def _verify_magic(ext: str, file_bytes: bytes) -> bool:
    """İçeriğin uzantıyla beklenen magic-byte imzasına sahip olup olmadığını doğrular."""
    signatures = _MAGIC_SIGNATURES.get(ext)
    if not signatures:
        return False
    head = file_bytes[:16]
    return any(head.startswith(sig) for sig in signatures)


def parse_file(filename: str, file_bytes: bytes) -> str:
    ext = Path(filename).suffix.lower()

    if not _verify_magic(ext, file_bytes):
        raise ValueError(
            f"Dosya içeriği {ext} uzantısıyla uyuşmuyor "
            "(magic-byte doğrulaması başarısız)."
        )

    if ext == ".pdf":
        return parse_pdf(file_bytes)
    elif ext in [".jpg", ".jpeg", ".png", ".bmp", ".tiff"]:
        return parse_image(file_bytes)
    elif ext == ".docx":
        return parse_docx(file_bytes)
    elif ext == ".xlsx":
        return parse_xlsx(file_bytes)
    else:
        raise ValueError(f"Desteklenmeyen dosya tipi: {ext}")